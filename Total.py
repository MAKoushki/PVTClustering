import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
from openpyxl import Workbook

data = pd.read_excel("data clustering.xlsx")
data.head()

#data.plot(kind='scatter', x='No k', y='Nw k')
#plt.show()

X = data.iloc[:,[6,7,8,9,10,11]].values

#elbow method
wcss = []
for i in range(1,11):
    k_means = KMeans(n_clusters=i,init='k-means++', random_state=42)
    k_means.fit(X)
    wcss.append(k_means.inertia_)
#plot elbow curve
plt.plot(np.arange(1,11),wcss)
plt.xlabel('Clusters')
plt.ylabel('SSE')
plt.show()

kmeans = KMeans(n_clusters = 4, init = 'k-means++',  random_state=42)
y_kmeans = kmeans.fit_predict(X)
#
wb = Workbook()
sh = wb.active
sh.title = "Sheet"   # sheet name
sh['F1'] = 'Rocktype'

#A=y_kmeans == 0
#B=y_kmeans == 1
#C=y_kmeans == 2
#D=y_kmeans == 3


#for itr in range(0,len(A)):
#    if (A[itr]==True):
#        sh.cell(row=itr+2, column=6).value=0
#    elif(B[itr]==True):
#        sh.cell(row=itr+2, column=6).value=1
#    elif(C[itr]==True):
#        sh.cell(row=itr+2, column=6).value=2
#    elif(D[itr]==True):
#        sh.cell(row=itr+2, column=6).value=3
for itr in range(0,len(y_kmeans)):
    sh.cell(row=itr+2, column=6).value=(y_kmeans[itr]+1)


wb.save("RockType.xlsx")

ANo=X[y_kmeans == 0, 2]
ANw=X[y_kmeans == 0, 3]
BNo=X[y_kmeans == 1, 2]
BNw=X[y_kmeans == 1, 3]
CNo=X[y_kmeans == 2, 2]
CNw=X[y_kmeans == 2, 3]
DNo=X[y_kmeans == 3, 2]
DNw=X[y_kmeans == 3, 3]

AKro=X[y_kmeans == 0, 4]
AKrw=X[y_kmeans == 0, 5]
BKro=X[y_kmeans == 1, 4]
BKrw=X[y_kmeans == 1, 5]
CKro=X[y_kmeans == 2, 4]
CKrw=X[y_kmeans == 2, 5]
DKro=X[y_kmeans == 3, 4]
DKrw=X[y_kmeans == 3, 5]

ASwr=X[y_kmeans == 0, 0]
ASor=X[y_kmeans == 0, 1]
BSwr=X[y_kmeans == 1, 0]
BSor=X[y_kmeans == 1, 1]
CSwr=X[y_kmeans == 2, 0]
CSor=X[y_kmeans == 2, 1]
DSwr=X[y_kmeans == 3, 0]
DSor=X[y_kmeans == 3, 1]

MeanANo= (np.sum(ANo))/len(ANo)
MeanANw= (np.sum(ANw))/len(ANw)
MeanBNo= (np.sum(BNo))/len(BNo)
MeanBNw= (np.sum(BNw))/len(BNw)
MeanCNo= (np.sum(CNo))/len(CNo)
MeanCNw= (np.sum(CNw))/len(CNw)
MeanDNo= (np.sum(DNo))/len(DNo)
MeanDNw= (np.sum(DNw))/len(DNw)

MeanAKro= (np.sum(AKro))/len(AKro)
MeanAKrw= (np.sum(AKrw))/len(AKrw)
MeanBKro= (np.sum(BKro))/len(BKro)
MeanBKrw= (np.sum(BKrw))/len(BKrw)
MeanCKro= (np.sum(CKro))/len(CKro)
MeanCKrw= (np.sum(CKrw))/len(CKrw)
MeanDKro= (np.sum(DKro))/len(DKro)
MeanDKrw= (np.sum(DKrw))/len(DKrw)

MeanASwr= (np.sum(ASwr))/len(ASwr)
MeanASor= (np.sum(ASor))/len(ASor)
MeanBSwr= (np.sum(BSwr))/len(BSwr)
MeanBSor= (np.sum(BSor))/len(BSor)
MeanCSwr= (np.sum(CSwr))/len(CSwr)
MeanCSor= (np.sum(CSor))/len(CSor)
MeanDSwr= (np.sum(DSwr))/len(DSwr)
MeanDSor= (np.sum(DSor))/len(DSor)

StdANo=np.std(ANo)
StdANw=np.std(ANw)
StdBNo=np.std(BNo)
StdBNw=np.std(BNw)
StdCNo=np.std(CNo)
StdCNw=np.std(CNw)
StdDNo=np.std(DNo)
StdDNw=np.std(DNw)

StdAKro=np.std(AKro)
StdAKrw=np.std(AKrw)
StdBKro=np.std(BKro)
StdBKrw=np.std(BKrw)
StdCKro=np.std(CKro)
StdCKrw=np.std(CKrw)
StdDKro=np.std(DKro)
StdDKrw=np.std(DKrw)

StdASwr=np.std(ASwr)
StdASor=np.std(ASor)
StdBSwr=np.std(BSwr)
StdBSor=np.std(BSor)
StdCSwr=np.std(CSwr)
StdCSor=np.std(CSor)
StdDSwr=np.std(DSwr)
StdDSor=np.std(DSor)

wb = Workbook()
sh = wb.active
sh.title = "Sheet2"   # sheet name

sh['B1'] = 'Cluster1'
sh['C1'] = 'Cluster2'
sh['D1'] = 'Cluster3'
sh['E1'] = 'Cluster4'
sh['A2'] = 'MeanNo'
sh['A3'] = 'MeanNw'
sh['A4'] = 'MeanSwc'
sh['A5'] = 'MeanSor'
sh['A6'] = 'MeanKromax'
sh['A7'] = 'MeanKrwmax'

sh['A8'] = 'StdNw'
sh['A9'] = 'StdNo'
sh['A10'] = 'StdSwc'
sh['A11'] = 'StdSor'
sh['A12'] = 'StdKromax'
sh['A13'] = 'StdKrwmax'


sh.cell(row=2, column=2).value  = MeanANo
sh.cell(row=2, column=3).value  = MeanBNo
sh.cell(row=2, column=4).value  = MeanCNo
sh.cell(row=2, column=5).value  = MeanDNo

sh.cell(row=3, column=2).value  = MeanANw
sh.cell(row=3, column=3).value  = MeanBNw
sh.cell(row=3, column=4).value  = MeanCNw
sh.cell(row=3, column=5).value  = MeanDNw

sh.cell(row=4, column=2).value  = MeanASwr
sh.cell(row=4, column=3).value  = MeanBSwr
sh.cell(row=4, column=4).value  = MeanCSwr
sh.cell(row=4, column=5).value  = MeanDSwr

sh.cell(row=5, column=2).value  = MeanASor
sh.cell(row=5, column=3).value  = MeanBSor
sh.cell(row=5, column=4).value  = MeanCSor
sh.cell(row=5, column=5).value  = MeanDSor

sh.cell(row=6, column=2).value  = MeanAKro
sh.cell(row=6, column=3).value  = MeanBKro
sh.cell(row=6, column=4).value  = MeanCKro
sh.cell(row=6, column=5).value  = MeanDKro

sh.cell(row=7, column=2).value  = MeanAKrw
sh.cell(row=7, column=3).value  = MeanBKrw
sh.cell(row=7, column=4).value  = MeanCKrw
sh.cell(row=7, column=5).value  = MeanDKrw


sh.cell(row=8, column=2).value  = StdANw
sh.cell(row=8, column=3).value  = StdBNw
sh.cell(row=8, column=4).value  = StdCNw
sh.cell(row=8, column=5).value  = StdDNw

sh.cell(row=9, column=2).value  = StdANo
sh.cell(row=9, column=3).value  = StdBNo
sh.cell(row=9, column=4).value  = StdCNo
sh.cell(row=9, column=5).value  = StdDNo

sh.cell(row=10, column=2).value  = StdASwr
sh.cell(row=10, column=3).value  = StdBSwr
sh.cell(row=10, column=4).value  = StdCSwr
sh.cell(row=10, column=5).value  = StdDSwr

sh.cell(row=11, column=2).value  = StdASor
sh.cell(row=11, column=3).value  = StdBSor
sh.cell(row=11, column=4).value  = StdCSor
sh.cell(row=11, column=5).value  = StdDSor

sh.cell(row=12, column=2).value  = StdAKro
sh.cell(row=12, column=3).value  = StdBKro
sh.cell(row=12, column=4).value  = StdCKro
sh.cell(row=12, column=5).value  = StdDKro

sh.cell(row=13, column=2).value  = StdAKrw
sh.cell(row=13, column=3).value  = StdBKrw
sh.cell(row=13, column=4).value  = StdCKrw
sh.cell(row=13, column=5).value  = StdDKrw


wb.save("ResultOridinaryKriging.xlsx")

plt.scatter(X[y_kmeans == 0, 2],
            X[y_kmeans == 0, 3],
            s=8, color='red', label='Cluster1')

plt.scatter(X[y_kmeans == 1, 2],
            X[y_kmeans == 1, 3],

            s = 4, c='blue', label='Cluster2')

plt.scatter(X[y_kmeans == 2, 2],
            X[y_kmeans == 2,3],

            s = 4, c='green', label='Cluster3')

plt.scatter(X[y_kmeans == 3, 2],
            X[y_kmeans == 3, 3],

            s = 4, c='cyan', label='Cluster4')

plt.scatter(kmeans.cluster_centers_[:,0],kmeans.cluster_centers_[:,1],
            s=50,c='yellow',label='center')
plt.legend()
plt.xlabel('Nw')
plt.ylabel('No')
plt.show()